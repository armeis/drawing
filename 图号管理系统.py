import tkinter
import tkinter.messagebox
import tkinter.ttk
import pymysql
import configparser
import windnd
import os
import openpyxl
from datetime import datetime
#import datetime
from tkinter import filedialog
from ast import literal_eval
import time
import ctypes
from tkcalendar import Calendar

is_admin=False  # 登录状态
current_dir=os.path.dirname(os.path.abspath(__file__))  # 获取当前目录
config=configparser.ConfigParser()  # 实例化ConfigParser
config.read(os.path.join(current_dir,"config.ini"))   # 读取当前目录下的config.ini文件
servername=config.get("DATABASE","host")    # 读取 [DATABASE] 分组下的 host 的值
serverdatabase=config.get("DATABASE","database")
serveruser=config.get("DATABASE","username")
serverpassword=config.get("DATABASE","password")
export_path=config.get("PATH","export_path")
main_path=config.get("PATH","main_path")

myappid="江西晖新材料有限公司图纸图号管理系统 V1.0"
ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)      # 设置任务栏图标和程序图标一致
# 获取数据库中图纸分类并去重将结果组合成列表
classify_sql = "select distinct drawing_classification from blueprint"
classify_list=[]
try:
    conn_classify = pymysql.connect(host=servername, user=serveruser, password=serverpassword, database=serverdatabase)
    cursor_classify = conn_classify.cursor()
    cursor_classify.execute(classify_sql)
    class_name = cursor_classify.fetchall()
    cursor_classify.close()
    conn_classify.close()
    for i in range(len(class_name)):
        classify_list.append(class_name[i][0])
except:
    classify_list=["切割图纸","水刀图纸","客供原稿"]    # 图纸分类列表

def data_query(self,allsql,chaxun_sql,orderby):
    chaxun = ""
    if self.itemNo.get() != "":
        chaxun = chaxun + "partnumber.item_No like '%" + self.itemNo.get() + "%'"
    if self.item_name.get() != "":
        if chaxun == "":
            chaxun = chaxun + "partnumber.item_name like '%" + self.item_name.get() + "%'"
        else:
            chaxun = chaxun + " and partnumber.item_name like '%" + self.item_name.get() + "%'"
    if self.item_specs.get() != "":
        if chaxun == "":
            chaxun = chaxun + "partnumber.item_specs like '%" + self.item_specs.get() + "%'"
        else:
            chaxun = chaxun + " and partnumber.item_specs like '%" + self.item_specs.get() + "%'"
    if self.drawing_name.get() != "":
        if chaxun == "":
            chaxun = chaxun + "blueprint.drawing_name like '%" + self.drawing_name.get() + "%'"
        else:
            chaxun = chaxun + " and blueprint.drawing_name like '%" + self.drawing_name.get() + "%'"
    if self.drawing_code.get() != "":
        if chaxun == "":
            chaxun = chaxun + "blueprint.drawing_code like '%" + self.drawing_code.get() + "%'"
        else:
            chaxun = chaxun + " and blueprint.drawing_code like '%" + self.drawing_code.get() + "%'"
    if self.drawing_identifier.get() != "":
        if chaxun == "":
            chaxun = chaxun + "blueprint.drawing_identifier like '%" + self.drawing_identifier.get() + "%'"
        else:
            chaxun = chaxun + " and blueprint.drawing_identifie like '%" + self.drawing_identifier.get() + "%'"
    if self.drawing_parameter.get() != "":
        if chaxun == "":
            chaxun = chaxun + "blueprint.drawing_parameter like '%" + self.drawing_parameter.get() + "%'"
        else:
            chaxun = chaxun + " and blueprint.drawing_parameter like '%" + self.drawing_parameter.get() + "%'"
    if self.drawing_edition.get() != "":
        if chaxun == "":
            chaxun = chaxun + "blueprint.drawing_edition='" + self.drawing_edition.get() + "'"
        else:
            chaxun = chaxun + " and blueprint.drawing_edition='" + self.drawing_edition.get() + "'"
    if self.drawing_classification.get() != "":
        if chaxun == "":
            chaxun = chaxun + "blueprint.drawing_classification like '%" + self.drawing_classification.get() + "%'"
        else:
            chaxun = chaxun + " and blueprint.drawing_classification like '%" + self.drawing_classification.get() + "%'"
    if self.document_name.get() != "":
        if chaxun == "":
            chaxun = chaxun + "blueprint.document_name like '%" + self.document_name.get() + "%'"
        else:
            chaxun = chaxun + " and blueprint.document_name like '%" + self.document_name.get() + "%'"
    if self.document_type.get() != "":
        if chaxun == "":
            chaxun = chaxun + "blueprint.document_type='" + self.document_type.get() + "'"
        else:
            chaxun = chaxun + " and blueprint.document_type='" + self.document_type.get() + "'"
    if self.document_path.get() != "":
        if chaxun == "":
            chaxun = chaxun + "blueprint.document_path like '%" + self.document_path.get() + "%'"
        else:
            chaxun = chaxun + " and blueprint.document_path like '%" + self.document_path.get() + "%'"
    if self.drawing_remarks.get() != "":
        if chaxun == "":
            chaxun = chaxun + "blueprint.drawing_remarks like '%" + self.drawing_remarks.get() + "%'"
        else:
            chaxun = chaxun + " and blueprint.drawing_remarks like '%" + self.drawing_remarks.get() + "%'"
    if chaxun != "":
        chaxun = chaxun_sql + " where " + chaxun + orderby
    else:
        chaxun = allsql
    return chaxun

class Welcome():

    def __init__(self,master):

        self.master= master
        # screenheight 屏幕高度
        screenheight = master.winfo_screenheight()/2
        # screenwidth 屏幕宽度
        screenwidth = master.winfo_screenwidth()/2
        tkwidth=700
        tkheight=538
        self.master.geometry(f"{tkwidth}x{tkheight}")
        self.master.geometry(f"+{int(screenwidth-tkwidth/2)}+{int(screenheight-tkheight/2)}")   # 窗口移动到屏幕正中间
        self.master.title("江西晖新材料有限公司图纸图号管理系统 V1.0")
        self.master.iconphoto(True,tkinter.PhotoImage(file=os.path.join(current_dir,"Siltherm.png")))     # 自定义程序图标
        self.master.resizable(False,False)  # 禁止窗体最大化

        imagefilename = tkinter.PhotoImage(file=os.path.join(current_dir,"背景1.gif"))
        background_label=tkinter.Label(self.master, text="江西晖新材料有限公司图纸图号管理系统",font=("黑体",25),image=imagefilename,compound="bottom")
        background_label.place(x=0,y=0,relwidth=1,relheight=1)
        background_label.image= imagefilename

        self.button1=tkinter.Button(self.master,text="管理员登录",bg="#f0f0f0",font=("宋体",18),command=self.gotoRegister)
        self.button1.place(x=275,y=200,width=150,height=40)
        self.button3 = tkinter.Button(self.master, text="直接进入", bg="#f0f0f0", font=("宋体", 18), command=self.gotoWages)
        self.button3.place(x=275, y=250, width=150, height=40)
        self.button2=tkinter.Button(self.master,text="退  出",bg="#f0f0f0",font=("宋体",18),command=self.finish)
        self.button2.place(x=275,y=300,width=150,height=40)

    def finish(self):

        self.master.destroy()

    def gotoRegister(self):

        root2=tkinter.Toplevel(self.master)
        self.master.withdraw()  # 窗口变得不可见
        myGUI=Register(root2,self.master)      # 添加解决方案返回主窗口

    def gotoWages(self):

        root2=tkinter.Toplevel(self.master)
        self.master.withdraw()  # 窗口变得不可见
        # myGUI=VisitorDrawingNo(root2,self.master)      # 添加解决方案返回主窗口
        myGUI = DrawingNo(root2, self.master)  # 添加解决方案返回主窗口

class Register():

    def __init__(self,master,mainwnd):

        self.account= tkinter.StringVar()
        self.password= tkinter.StringVar()

        self.mainwnd = mainwnd  # 存储主窗口
        self.register_master= master
        # screenheight 屏幕高度
        screenheight = master.winfo_screenheight() / 2
        # screenwidth 屏幕宽度
        screenwidth = master.winfo_screenwidth() / 2
        tkwidth = 500
        tkheight = 300
        self.register_master.geometry(f"{tkwidth}x{tkheight}")
        self.register_master.geometry(f"+{int(screenwidth - tkwidth / 2)}+{int(screenheight - tkheight / 2)}")  # 窗口移动到屏幕正中间
        self.register_master.attributes("-topmost",True)    # 窗口始终置项
        self.register_master.title("用户登录")
        self.register_master.after(1,lambda:self.register_master.focus_force()) # 窗口获取焦点

        self.label1=tkinter.Label(self.register_master,text="用户名：",fg="black",font=("宋体",25),width=12,anchor="e").grid(row=0,column=0,pady=50)
        self.label2=tkinter.Label(self.register_master,text="密  码：",font=("宋体",25),width=12,anchor="e").grid(row=1,column=0)
        self.label3=tkinter.Label(self.register_master,text="").grid(row=2,column=0)

        self.account1= tkinter.Entry(self.register_master,textvariable=self.account,width=30)
        self.account1.grid(row=0, column=1,ipady=8)
        self.account1.focus_set()
        self.password1= tkinter.Entry(self.register_master,textvariable=self.password,width=30,show="*")
        self.password1.grid(row=1, column=1,ipady=8)
        self.button1=tkinter.Button(self.register_master,text="确定",fg="blue",font=("黑体",18),command=self.Sign_in)
        self.button1.grid(row=3,column=0,ipadx=30,ipady=6)
        self.register_master.bind("<Return>", lambda e:self.Sign_in())      # 将确定按钮绑定到回车键
        self.button2=tkinter.Button(self.register_master,text="关闭",fg="blue",font=("黑体",18),command=self.myquit)
        self.button2.grid(row=3,column=1,ipadx=30,ipady=6)

        self.register_master.protocol("WM_DELETE_WINDOW",self.myquit)    # 点击关闭X实现关闭并返回主窗口

    def myquit(self):
        self.register_master.destroy()  # 关闭当前窗口
        self.mainwnd.update()  # 更新主窗口
        self.mainwnd.deiconify()  # 取消最小化主窗口
    def Sign_in(self):
        global is_admin
        if self.account.get()!="" or self.password.get()!="":
            name_sql = f"select * from user where name='{self.account.get()}'"
            conn = pymysql.connect(host=servername, user=serveruser, password=serverpassword, database=serverdatabase)
            cursor_1 = conn.cursor()
            cursor_1.execute(name_sql)
            name = cursor_1.fetchall()
            cursor_1.close()
            conn.close()
            try:
                if self.account.get()==name[0][0] and self.password.get()==name[0][1]:
                    is_admin=True
                    # admin_true()
                    self.register_master.destroy()  # 关闭当前窗口
                    self.mainwnd.update()  # 更新主窗口
                    self.mainwnd.deiconify()  # 取消最小化主窗口
                    root1 = tkinter.Toplevel(self.mainwnd)
                    self.mainwnd.withdraw()  # 窗口变得不可见
                    myGUI = DrawingNo(root1, self.register_master)  # 添加解决方案返回主窗口
                else:
                    tkinter.messagebox.showerror(parent=self.register_master,title="错误", message="请输入正确的用户名和密码。")
            except:
                tkinter.messagebox.showerror(parent=self.register_master,title="错误",message="请输入正确的用户名和密码。")
        else:
            tkinter.messagebox.showerror(parent=self.register_master,title="错误",message="用户名和密码不能为空。")

class DrawingNo():
    global is_admin
    def __init__(self,master,mainwnd):
        titles=["ID","料号","品名","规格","图纸ID","图纸名称","图纸编码","图纸编号","图纸参数","图纸版本","图纸分类","创建时间","文件名称","文件类型","文件路径","图纸备注"]
        self.itemNo= tkinter.StringVar()
        self.item_name=tkinter.StringVar()
        self.item_specs=tkinter.StringVar()
        self.drawing_name= tkinter.StringVar()
        self.drawing_identifier = tkinter.StringVar()
        self.drawing_parameter = tkinter.StringVar()
        self.drawing_edition = tkinter.StringVar()
        self.drawing_classification = tkinter.StringVar()
        self.document_name = tkinter.StringVar()
        self.document_type = tkinter.StringVar()
        self.document_path= tkinter.StringVar()
        self.drawing_remarks = tkinter.StringVar()
        self.drawing_code = tkinter.StringVar()

        self.mainwnd = mainwnd  # 存储主窗口
        self.master1= master
        self.master1.protocol("WM_DELETE_WINDOW", self.DrawingNoquit)  # 点击关闭X实现关闭并返回主窗口
        # screenheight 屏幕高度
        screenheight = master.winfo_screenheight() / 2
        # screenwidth 屏幕宽度
        screenwidth = master.winfo_screenwidth() / 2
        tkwidth = int(screenwidth*2-400)
        tkheight = int(screenheight*2-200)
        self.master1.geometry(f"{tkwidth}x{tkheight}")
        self.master1.geometry(f"+{int(screenwidth - tkwidth / 2)}+{int(screenheight - tkheight / 2-35)}")  # 窗口移动到屏幕正中间
        self.master1.after(1,lambda:self.master1.focus_force()) # 窗口获取焦点
        self.master1.title("江西晖能新材料有限公司图号管理系统")
        self.master1.bind("<Configure>",self.chanage_window)    # 监测窗口事件，绑定检测窗口大小改变自宝义函数

        # 小部件大小随窗口的增加而增长
        for i in range(1):
            self.master1.grid_rowconfigure(i,weight=1)
            for j in range(12):
                if j%2==0:
                    tkinter.Label(self.master1,text="",width=10,borderwidth=0,relief="sunken").grid(row=i,column=j,sticky="new")
                    self.master1.grid_columnconfigure(j,weight=1)
                else:
                    tkinter.Label(self.master1, text="", width=10, borderwidth=0, relief="sunken").grid(row=i, column=j,sticky="new")
                    self.master1.grid_columnconfigure(j, weight=2)
        self.label1=tkinter.Label(self.master1,text="料号",borderwidth=0,relief="sunken").grid(row=0,column=0)
        self.label12 = tkinter.Label(self.master1, text="品名").grid(row=0, column=2)
        self.label13 = tkinter.Label(self.master1, text="规格").grid(row=0, column=4)
        self.label2=tkinter.Label(self.master1,text="图纸名称").grid(row=0,column=6)
        self.label3=tkinter.Label(self.master1,text="图纸编号").grid(row=0,column=8)
        self.label4 = tkinter.Label(self.master1, text="图纸参数").grid(row=0, column=10)
        self.label5 = tkinter.Label(self.master1, text="图纸版本").grid(row=1, column=0)
        self.label6 = tkinter.Label(self.master1, text="图纸分类").grid(row=1, column=2)
        self.label7 = tkinter.Label(self.master1, text="文件名称").grid(row=1, column=4)
        self.label8 = tkinter.Label(self.master1, text="文件类型").grid(row=1, column=6)
        self.label9 = tkinter.Label(self.master1, text="文件路径").grid(row=1, column=8)
        self.label10 = tkinter.Label(self.master1, text="图纸备注").grid(row=1, column=10)
        self.label11 = tkinter.Label(self.master1, text="").grid(row=2, column=0)
        self.label14 = tkinter.Label(self.master1, text="",borderwidth=0,relief="sunken").grid(row=0, column=12)
        self.record_label = tkinter.Label(self.master1, text="查询到0条记录", borderwidth=1, relief="sunken")
        self.record_label.grid(row=156, column=11)

        self.text_itemNo= tkinter.Entry(self.master1, textvariable= self.itemNo,width=25)
        self.text_itemNo.grid(row=0, column=1)
        self.text_drawing_name= tkinter.Entry(self.master1, textvariable= self.drawing_name,width=25)
        self.text_drawing_name.grid(row=0, column=7)
        self.text_drawing_identifier = tkinter.Entry(self.master1, textvariable=self.drawing_identifier,width=25)
        self.text_drawing_identifier.grid(row=0, column=9)
        self.text_drawing_parameter = tkinter.Entry(self.master1, textvariable=self.drawing_parameter,width=25)
        self.text_drawing_parameter.grid(row=0, column=11)
        self.text_drawing_edition = tkinter.Entry(self.master1, textvariable=self.drawing_edition,width=25)
        self.text_drawing_edition.grid(row=1, column=1)
        self.text_drawing_classification = tkinter.ttk.Combobox(self.master1, textvariable=self.drawing_classification,values=classify_list,width=24)
        self.text_drawing_classification.grid(row=1, column=3)
        self.text_document_name = tkinter.Entry(self.master1, textvariable=self.document_name,width=25)
        self.text_document_name.grid(row=1, column=5)
        self.text_document_type = tkinter.ttk.Combobox(self.master1, textvariable=self.document_type,values=literal_eval(config.get("MULTIPLE","file_type")),width=24)
        self.text_document_type.grid(row=1, column=7)
        self.text_document_path = tkinter.Entry(self.master1, textvariable=self.document_path,width=25)
        self.text_document_path.grid(row=1, column=9)
        self.text_drawing_remarks = tkinter.Entry(self.master1, textvariable=self.drawing_remarks,width=25)
        self.text_drawing_remarks.grid(row=1, column=11)
        self.text_item_name = tkinter.Entry(self.master1, textvariable=self.item_name,width=25)
        self.text_item_name.grid(row=0, column=3)
        self.text_item_specs = tkinter.Entry(self.master1, textvariable=self.item_specs,width=25)
        self.text_item_specs.grid(row=0, column=5)

        self.button1=tkinter.Button(self.master1,text="查询料号对应图纸",fg="blue",command=self.all_query).grid(row=156,column=1)
        self.button2 = tkinter.Button(self.master1, text="查询图纸", fg="blue", command=self.blueprint_query).grid(row=156, column=2)
        self.button3 = tkinter.Button(self.master1, text="清空查找条件", fg="red",command=self.clear_find).grid(row=156, column=3)
        self.button4 = tkinter.Button(self.master1, text="修改选中行记录", fg="purple", command=self.select_modify)
        self.button4.grid(row=156, column=4)
        if is_admin:
            self.button4.config(state="normal")
        else:
            self.button4.config(state="disabled")
        self.button5 = tkinter.Button(self.master1, text="新增/插入记录", fg="green",command=self.insert_data)
        self.button5.grid(row=156, column=5)
        if is_admin:
            self.button5.config(state="normal")
        else:
            self.button5.config(state="disabled")
        self.button6 = tkinter.Button(self.master1, text="批量新增数据", fg="green",command=self.batchInsert)
        self.button6.grid(row=156, column=6)
        if is_admin:
            self.button6.config(state="normal")
        else:
            self.button6.config(state="disabled")
        self.button7 = tkinter.Button(self.master1, text="删除选中料号表记录", fg="red",command=self.deleteItem)
        self.button7.grid(row=156, column=7)
        if is_admin:
            self.button7.config(state="normal")
        else:
            self.button7.config(state="disabled")
        self.button8 = tkinter.Button(self.master1, text="删除选中图纸图号表记录", fg="red",command=self.deleteDrawing)
        self.button8.grid(row=156, column=8)
        if is_admin:
            self.button8.config(state="normal")
        else:
            self.button8.config(state="disabled")
        self.button9 = tkinter.Button(self.master1, text="导出查询结果", fg="blue",command=self.exportExcel).grid(row=156, column=9)
        self.button10=tkinter.Button(self.master1,text="退  出",fg="black",command=self.DrawingNoquit).grid(row=156,column=10,ipadx=25)

        self.last_time = time.time()
        self.master1.update()
        self.data_table=tkinter.ttk.Treeview(self.master1,columns=titles,height=int(round(self.master1.winfo_height()/22,0))-4,show="headings")
        treeview_column_width=[40,90,100,150,50,200,60,60,100,60,60,120,200,60,400,200]
        for i in range(16):
            self.data_table.column(titles[i],width=treeview_column_width[i])
            self.data_table.heading(titles[i],text=titles[i])
        self.data_table.grid(row=3,column=0,rowspan=150,columnspan=12,sticky="nsew")
        # 定义竖向滚动条
        self.vbar=tkinter.ttk.Scrollbar(self.master1,orient="vertical",command=self.data_table.yview)
        self.data_table.configure(yscrollcommand=self.vbar.set)
        self.vbar.grid(row=3,column=13,rowspan=150,sticky="NS")
        # 定义横向滚动条
        self.hbar = tkinter.ttk.Scrollbar(self.master1, orient="horizontal", command=self.data_table.xview)
        self.data_table.configure(xscrollcommand=self.hbar.set)
        self.hbar.grid(row=154, column=0, sticky="WE", columnspan=12)
        self.data_table.bind("<Button-3>",self.rightClickMenu)  # 绑定右键点击事件

        def treeview_sort_column(tv, col, reverse):  # Treeview、列名、排列方式
            l = [(tv.set(k, col), k) for k in tv.get_children('')]
            l.sort(reverse=reverse)  # 排序方式
            for index, (val, k) in enumerate(l):  # 根据排序后索引移动
                tv.move(k, '', index)
            tv.heading(col, command=lambda: treeview_sort_column(tv, col, not reverse))  # 重写标题，使之成为再点倒序的标题

        for col in self.data_table["columns"]:  # 给所有标题加（循环上边的“手工”）
            self.data_table.heading(col, text=col, command=lambda _col=col: treeview_sort_column(self.data_table, _col, False))

    def clear_find(self):   # 清空查找输入框
        self.text_item_specs.delete(0, tkinter.END)
        self.text_drawing_parameter.delete(0, tkinter.END)
        self.text_drawing_identifier.delete(0, tkinter.END)
        self.text_drawing_classification.delete(0, tkinter.END)
        self.text_drawing_name.delete(0, tkinter.END)
        self.text_document_type.delete(0, tkinter.END)
        self.text_document_path.delete(0, tkinter.END)
        self.text_document_name.delete(0, tkinter.END)
        self.text_drawing_edition.delete(0, tkinter.END)
        self.text_drawing_remarks.delete(0, tkinter.END)
        self.text_item_name.delete(0, tkinter.END)
        self.text_itemNo.delete(0, tkinter.END)

    def chanage_window(self, event=None):   # 更新窗口大小变动
        current_time=time.time()
        if(current_time-self.last_time)>0.5:
            self.master1.update()
            tree_height=self.master1.winfo_height()
            self.data_table.configure(height=int(round(tree_height/22,0))-4)

        self.last_time = time.time()

    def rightClickMenu(self,event):     # treeview右键菜单
        global is_admin
        def openfile():
            os.startfile(f"{self.data_table.item(rowID)['values'][14]}")
        def openfilepath():
            path=self.data_table.item(rowID)['values'][14][:len(self.data_table.item(rowID)['values'][14])-self.data_table.item(rowID)['values'][14][::-1].index("/")]
            os.startfile(path)
        rowID=self.data_table.identify("item",event.x,event.y)
        if rowID:
            menu=tkinter.Menu(self.master1,tearoff=0)
            if is_admin:
                menu.add_command(label="修改选中项数据",command=self.select_modify)
                menu.add_command(label="删除选中料号行", command=self.deleteItem)
                menu.add_command(label="删除选中图号行", command=self.deleteDrawing)
            menu.add_command(label="打  开  图  纸",command=openfile)
            menu.add_command(label="打开图纸文件夹", command=openfilepath)
            self.data_table.selection_set(rowID)
            self.data_table.focus_set()
            self.data_table.focus(rowID)
            menu.post(event.x_root,event.y_root)
        else:
            pass

    def all_query(self):    # 查询按钮
        allsql = "select partnumber.item_id,partnumber.item_No,partnumber.item_name,partnumber.item_specs,blueprint.drawing_id,blueprint.drawing_name,blueprint.drawing_code,blueprint.drawing_identifier,blueprint.drawing_parameter,blueprint.drawing_edition,blueprint.drawing_classification,blueprint.orderDate,blueprint.document_name,blueprint.document_type,blueprint.document_path,blueprint.drawing_remarks from partnumber inner join blueprint on partnumber.blueprint_id=blueprint.drawing_id order by partnumber.item_id desc"
        chaxun_sql = "select partnumber.item_id,partnumber.item_No,partnumber.item_name,partnumber.item_specs,blueprint.drawing_id,blueprint.drawing_name,blueprint.drawing_code,blueprint.drawing_identifier,blueprint.drawing_parameter,blueprint.drawing_edition,blueprint.drawing_classification,blueprint.orderDate,blueprint.document_name,blueprint.document_type,blueprint.document_path,blueprint.drawing_remarks from partnumber inner join blueprint on partnumber.blueprint_id=blueprint.drawing_id"
        orderby=" order by partnumber.item_id desc"
        chaxun=data_query(self,allsql, chaxun_sql,orderby)
        try:
            conn=pymysql.connect(host=servername,user=serveruser,password=serverpassword,database=serverdatabase)
            # 游标使用注意事项
            # 一个连接一次只能有一个游标的查询处于活跃状态，如下：
            cursor_1=conn.cursor()
            cursor_1.execute(chaxun)
            allsql_list=cursor_1.fetchall()

            for row in self.data_table.get_children():
                self.data_table.delete(row)

            for i in range(len(allsql_list)):
                allsql_list_text=str(allsql_list[i])
                self.data_table.insert("",0,text=allsql_list_text,values=allsql_list[i])

            cursor_1.close()
            conn.close()
            self.record_label.config(text=f"查询到{len(allsql_list)}条记录")
        except:
            tkinter.messagebox.showerror(title="错误", message="查找条件错误！\n请修改查找条件！")

    def blueprint_query(self):    # 查询按钮
        allsql = "select drawing_id,drawing_name,drawing_code,drawing_identifier,drawing_parameter,drawing_edition,drawing_classification,orderDate,document_name,document_type,document_path,drawing_remarks from blueprint order by drawing_id desc"
        chaxun_sql = "select drawing_id,drawing_name,drawing_code,drawing_identifier,drawing_parameter,drawing_edition,drawing_classification,orderDate,document_name,document_type,document_path,drawing_remarks from blueprint"
        orderby = " order by drawing_id desc"
        chaxun=data_query(self,allsql, chaxun_sql,orderby)
        try:
            conn=pymysql.connect(host=servername,user=serveruser,password=serverpassword,database=serverdatabase)
            # 游标使用注意事项
            # 一个连接一次只能有一个游标的查询处于活跃状态，如下：
            cursor_2=conn.cursor()
            cursor_2.execute(chaxun)
            allsql_list=cursor_2.fetchall()

            for row in self.data_table.get_children():
                self.data_table.delete(row)

            for i in range(len(allsql_list)):
                tablelist=allsql_list[i][:0]+("","","","")+allsql_list[i][0:]
                # print(type(tablelist))
                # print(tablelist)
                tablelist_text=str(tablelist)
                self.data_table.insert("",0,text=tablelist_text,values=tablelist)

            cursor_2.close()
            conn.close()
            self.record_label.config(text=f"查询到{len(allsql_list)}条记录")
        except:
            tkinter.messagebox.showerror(title="错误",message="只查询图纸时需要清空料号，\n品名和规格输入框的内容。")

    def drawingId_query(self,drawingId):
        drawing_sql = f"select drawing_id,drawing_name,drawing_code,drawing_identifier,drawing_parameter,drawing_edition,drawing_classification,orderDate,document_name,document_type,document_path,drawing_remarks from blueprint where drawing_id='{drawingId}'"
        conn = pymysql.connect(host=servername, user=serveruser, password=serverpassword, database=serverdatabase)
        cursor_4 = conn.cursor()
        cursor_4.execute(drawing_sql)
        drawing_list = cursor_4.fetchall()
        cursor_4.close()
        conn.close()
        return drawing_list

    def itemId_query(self,itemId):
        drawing_sql = f"select item_id,item_No,item_name,tem_specs,blueprint_id from partnumber where item_id='{itemId}'"
        conn = pymysql.connect(host=servername, user=serveruser, password=serverpassword, database=serverdatabase)
        cursor_5 = conn.cursor()
        cursor_5.execute(drawing_sql)
        item_list = cursor_5.fetchall()
        cursor_5.close()
        conn.close()
        return item_list

    def select_modify(self):
        curItem = self.data_table.focus()
        if curItem==None or curItem=="":
            tkinter.messagebox.showerror(title="错误",message="未选中任何行数据。")
        else:
            drawing_list = self.drawingId_query(drawingId=self.data_table.item(curItem)["values"][4])
            self.top1=tkinter.Toplevel()
            self.top1.title("修改选中记录")
            # screenheight 屏幕高度
            screenheight = self.top1.winfo_screenheight() / 2
            # screenwidth 屏幕宽度
            screenwidth = self.top1.winfo_screenwidth() / 2
            tkwidth = 800
            tkheight = 335
            self.top1.geometry(f"{tkwidth}x{tkheight}")
            self.top1.geometry(f"+{int(screenwidth - tkwidth / 2)}+{int(screenheight - tkheight / 2)}")
            self.top1.attributes("-topmost", True)  # 窗口始终置项

            self.top1.item_id = tkinter.IntVar(value=self.data_table.item(curItem)["values"][0])
            self.top1.item_No = tkinter.StringVar(value=self.data_table.item(curItem)["values"][1])
            self.top1.item_name = tkinter.StringVar(value=self.data_table.item(curItem)["values"][2])
            self.top1.item_specs = tkinter.StringVar(value=self.data_table.item(curItem)["values"][3])
            self.top1.blueprint_id = tkinter.IntVar(value=self.data_table.item(curItem)["values"][4])
            self.top1.drawing_id = tkinter.IntVar(value=drawing_list[0][0])
            self.top1.drawing_name = tkinter.StringVar(value=drawing_list[0][1])
            self.top1.drawing_code = tkinter.StringVar(value=drawing_list[0][2])
            self.top1.drawing_identifier = tkinter.StringVar(value=drawing_list[0][3])
            self.top1.drawing_parameter = tkinter.StringVar(value=drawing_list[0][4])
            self.top1.drawing_edition = tkinter.StringVar(value=drawing_list[0][5])
            self.top1.drawing_classification = tkinter.StringVar(value=drawing_list[0][6])
            self.top1.orderDate = tkinter.StringVar(value=drawing_list[0][7])
            self.top1.document_name = tkinter.StringVar(value=drawing_list[0][8])
            self.top1.document_type = tkinter.StringVar(value=drawing_list[0][9])
            self.top1.document_path = tkinter.StringVar(value=drawing_list[0][10])
            self.top1.drawing_remarks = tkinter.StringVar(value=drawing_list[0][11])

            self.Lable1=tkinter.Label(self.top1,text="料号表partnumber").grid(row=0,column=0,columnspan=4)
            self.Lable2 = tkinter.Label(self.top1, text="图纸表blueprint").grid(row=0, column=6, columnspan=5)
            self.Lable3 = tkinter.Label(self.top1, text="编号：").grid(row=1, column=0)
            self.Lable4 = tkinter.Label(self.top1, text="料号：").grid(row=2, column=0)
            self.Lable5 = tkinter.Label(self.top1, text="品名：").grid(row=3, column=0)
            self.Lable6 = tkinter.Label(self.top1, text="规格：").grid(row=4, column=0)
            self.Lable7 = tkinter.Label(self.top1, text="图纸ID：").grid(row=5, column=0)
            self.Lable8 = tkinter.Label(self.top1, text="编号：").grid(row=1, column=5)
            self.Lable9 = tkinter.Label(self.top1, text="图纸名称：").grid(row=2, column=5)
            self.Lable10 = tkinter.Label(self.top1, text="图纸编码：").grid(row=3, column=5)
            self.Lable11 = tkinter.Label(self.top1, text="图纸编号：").grid(row=4, column=5)
            self.Lable12 = tkinter.Label(self.top1, text="图纸参数：").grid(row=5, column=5)
            self.Lable13 = tkinter.Label(self.top1, text="图纸版本：").grid(row=6, column=5)
            self.Lable14 = tkinter.Label(self.top1, text="图纸分类：").grid(row=7, column=5)
            self.Lable15 = tkinter.Label(self.top1, text="创建时间：").grid(row=8, column=5)
            self.Lable16 = tkinter.Label(self.top1, text="文件名称：").grid(row=9, column=5)
            self.Lable17 = tkinter.Label(self.top1, text="文件类型：").grid(row=10, column=5)
            self.Lable18 = tkinter.Label(self.top1, text="文件路径：").grid(row=11, column=5)
            self.Lable19 = tkinter.Label(self.top1, text="图纸备注：").grid(row=12, column=5)
            self.Lable20 = tkinter.Label(self.top1, text="拖入图纸文件到此框中\n自动获取图纸文件名和文件路径", relief="groove", width=40,
                                         height=8, fg="purple")
            self.Lable20.grid(row=6, column=0, rowspan=7, columnspan=4)

            self.text1 = tkinter.Entry(self.top1, textvariable=self.top1.item_id,width=40)
            self.text1.grid(row=1, column=1,columnspan=3)
            self.text2 = tkinter.Entry(self.top1, textvariable=self.top1.item_No,width=40)
            self.text2.grid(row=2, column=1,columnspan=3)
            self.text3 = tkinter.Entry(self.top1, textvariable=self.top1.item_name,width=40)
            self.text3.grid(row=3, column=1,columnspan=3)
            self.text4 = tkinter.Entry(self.top1, textvariable=self.top1.item_specs,width=40)
            self.text4.grid(row=4, column=1,columnspan=3)
            self.text5 = tkinter.Entry(self.top1, textvariable=self.top1.blueprint_id,width=40)
            self.text5.grid(row=5, column=1,columnspan=3)
            self.text6 = tkinter.Entry(self.top1, textvariable=self.top1.drawing_id,width=55)
            self.text6.grid(row=1, column=6,columnspan=4)
            self.text7 = tkinter.Entry(self.top1, textvariable=self.top1.drawing_name,width=55)
            self.text7.grid(row=2, column=6,columnspan=4)
            self.text8 = tkinter.Entry(self.top1, textvariable=self.top1.drawing_code,width=55)
            self.text8.grid(row=3, column=6,columnspan=4)
            self.text9 = tkinter.Entry(self.top1, textvariable=self.top1.drawing_identifier,width=55)
            self.text9.grid(row=4, column=6,columnspan=4)
            self.text10 = tkinter.Entry(self.top1, textvariable=self.top1.drawing_parameter,width=55)
            self.text10.grid(row=5, column=6,columnspan=4)
            self.text11 = tkinter.Entry(self.top1, textvariable=self.top1.drawing_edition,width=55)
            self.text11.grid(row=6, column=6,columnspan=4)
            #self.text12 = tkinter.Entry(self.top1, textvariable=self.top1.drawing_classification,width=55)
            self.text12 = tkinter.ttk.Combobox(self.top1, textvariable=self.top1.drawing_classification, width=52,values=classify_list)
            self.text12.grid(row=7,column=6,columnspan=4)
            self.text13 = tkinter.Entry(self.top1, textvariable=self.top1.orderDate,width=55)
            self.text13.grid(row=8, column=6,columnspan=4)
            self.text14 = tkinter.Entry(self.top1, textvariable=self.top1.document_name,width=55)
            self.text14.grid(row=9,column=6,columnspan=4)
            self.text15 = tkinter.Entry(self.top1, textvariable=self.top1.document_type,width=55)
            self.text15.grid(row=10, column=6,columnspan=4)
            self.text16 = tkinter.Entry(self.top1, textvariable=self.top1.document_path,width=55)
            self.text16.grid(row=11, column=6,columnspan=4)
            self.text17 = tkinter.Entry(self.top1, textvariable=self.top1.drawing_remarks,width=55)
            self.text17.grid(row=12, column=6,columnspan=4)


            self.button1 = tkinter.Button(self.top1, text="修改料号表", fg="blue",width=15,command=self.modifyItem).grid(row=13,column=1)
            self.button1 = tkinter.Button(self.top1, text="修改图纸图号表", fg="blue",width=15,command=self.modifyblueprint).grid(row=13, column=6)
            self.button1 = tkinter.Button(self.top1, text="清空料号表文本框", fg="red",width=15,command=self.cleaitemNo).grid(row=13, column=2)
            self.button1 = tkinter.Button(self.top1, text="清空图号表文本框", fg="red", width=15,command=self.cleadrawingNo).grid(row=13, column=7)
            self.button1 = tkinter.Button(self.top1, text="退  出", fg="black",width=15,command=self.top1quit).grid(row=13, column=9)

            def dragged_files(files):  # 获取文件路径和文件名及扩展名
                fullpath = "\n".join((item.decode("gbk") for item in files)).replace("\\", "/")
                filetype = fullpath[len(fullpath) - fullpath[::-1].index("."):]
                filename = fullpath[
                           len(fullpath) - fullpath[::-1].index("/"):len(fullpath) - fullpath[::-1].index(".") - 1]
                new_path=fullpath.replace(main_path,"")
                self.top1.document_type.set(value=filetype)
                self.top1.document_path.set(value=fullpath)
                self.top1.document_name.set(value=filename)
                self.top1.drawing_name.set(value=filename)
                self.top1.drawing_classification.set(value=new_path[:new_path.index("/")])

            def calendar(event):    # 日历选择小部件
                def print_sel():
                    self.text13.delete(0, tkinter.END)
                    self.text13.insert(0, str(cal.selection_get()))
                    top_cal.destroy()

                x, y = event.widget.winfo_pointerxy()   # 获取鼠标坐标
                top_cal = tkinter.Toplevel(self.top1)
                w, h = top_cal.winfo_screenwidth(), top_cal.winfo_screenheight()    # 获取屏幕分辨率的长和宽
                if x + 250 > w:     # 如果鼠标X位置+小部件宽大于屏幕宽
                    x = w - 260
                if y + 200 > h-80:
                    y = h - 280
                top_cal.geometry(f"250x200+{str(x)}+{str(y)}")
                top_cal.attributes("-topmost", True)  # 窗口始终置项
                cal = Calendar(top_cal,locale="zh", selectmode='day',cursor="hand1", year=datetime.now().year, month=datetime.now().month,day=datetime.now().day)
                cal.pack(fill="both", expand=True)
                tkinter.Button(top_cal, text=" 确 定 ", command=print_sel).pack()

            self.button_cal = tkinter.Button(self.top1, text='v', command=calendar, width=1,bd=0)
            self.button_cal.grid(row=8,column=9,sticky="e")
            self.button_cal.bind("<Button-1>", calendar)

            windnd.hook_dropfiles(self.Lable20, func=dragged_files)

            self.top1.protocol("WM_DELETE_WINDOW", self.top1quit)  # 点击关闭X实现关闭并返回主窗口

    def insert_data(self):
        self.top2 = tkinter.Toplevel()
        self.master1.state("iconic")
        #myGUI = Register(root2, self.master)  # 添加解决方案返回主窗口
        self.top2.title("新增记录")
        # screenheight 屏幕高度
        screenheight = self.top2.winfo_screenheight() / 2
        # screenwidth 屏幕宽度
        screenwidth = self.top2.winfo_screenwidth() / 2
        tkwidth = 800
        tkheight = 335
        self.top2.geometry(f"{tkwidth}x{tkheight}")
        self.top2.geometry(f"+{int(screenwidth - tkwidth / 2)}+{int(screenheight - tkheight / 2)}")
        self.top2.attributes("-topmost", True)  # 窗口始终置项

        self.top2.item_id = tkinter.IntVar()
        self.top2.item_No = tkinter.StringVar()
        self.top2.item_name = tkinter.StringVar()
        self.top2.item_specs = tkinter.StringVar()
        self.top2.blueprint_id = tkinter.IntVar()
        self.top2.drawing_id = tkinter.IntVar()
        self.top2.drawing_name = tkinter.StringVar()
        self.top2.drawing_code = tkinter.StringVar()
        self.top2.drawing_identifier = tkinter.StringVar()
        self.top2.drawing_parameter = tkinter.StringVar()
        self.top2.drawing_edition = tkinter.StringVar()
        self.top2.drawing_classification = tkinter.StringVar()
        self.top2.orderDate = tkinter.StringVar()
        self.top2.document_name = tkinter.StringVar()
        self.top2.document_type = tkinter.StringVar()
        self.top2.document_path = tkinter.StringVar()
        self.top2.drawing_remarks = tkinter.StringVar()

        self.Lable1 = tkinter.Label(self.top2, text="料号表partnumber（先添加料号对应的图纸资料）").grid(row=0, column=0, columnspan=4)
        self.Lable2 = tkinter.Label(self.top2, text="图纸表blueprint").grid(row=0, column=6, columnspan=5)
        self.Lable3 = tkinter.Label(self.top2, text="编号：").grid(row=1, column=0)
        self.Lable4 = tkinter.Label(self.top2, text="料号：").grid(row=2, column=0)
        self.Lable5 = tkinter.Label(self.top2, text="品名：").grid(row=3, column=0)
        self.Lable6 = tkinter.Label(self.top2, text="规格：").grid(row=4, column=0)
        self.Lable7 = tkinter.Label(self.top2, text="图纸ID：").grid(row=5, column=0)
        self.Lable8 = tkinter.Label(self.top2, text="编号：").grid(row=1, column=5)
        self.Lable9 = tkinter.Label(self.top2, text="图纸名称：").grid(row=2, column=5)
        self.Lable10 = tkinter.Label(self.top2, text="图纸编码：").grid(row=3, column=5)
        self.Lable11 = tkinter.Label(self.top2, text="图纸编号：").grid(row=4, column=5)
        self.Lable12 = tkinter.Label(self.top2, text="图纸参数：").grid(row=5, column=5)
        self.Lable13 = tkinter.Label(self.top2, text="图纸版本：").grid(row=6, column=5)
        self.Lable14 = tkinter.Label(self.top2, text="图纸分类：").grid(row=7, column=5)
        self.Lable15 = tkinter.Label(self.top2, text="创建时间：").grid(row=8, column=5)
        self.Lable16 = tkinter.Label(self.top2, text="文件名称：").grid(row=9, column=5)
        self.Lable17 = tkinter.Label(self.top2, text="文件类型：").grid(row=10, column=5)
        self.Lable18 = tkinter.Label(self.top2, text="文件路径：").grid(row=11, column=5)
        self.Lable19 = tkinter.Label(self.top2, text="图纸备注：").grid(row=12, column=5)
        self.Lable20 = tkinter.Label(self.top2, text="拖入图纸文件到此框中\n自动获取图纸文件名和文件路径",relief="groove",width=40,height=8, fg="purple")
        self.Lable20.grid(row=6, column=0,rowspan=7,columnspan=4)

        self.text1 = tkinter.Entry(self.top2, textvariable=self.top2.item_id, width=40)
        self.text1.grid(row=1, column=1, columnspan=3)
        self.text2 = tkinter.Entry(self.top2, textvariable=self.top2.item_No, width=40)
        self.text2.grid(row=2, column=1, columnspan=3)
        self.text3 = tkinter.Entry(self.top2, textvariable=self.top2.item_name, width=40)
        self.text3.grid(row=3, column=1, columnspan=3)
        self.text4 = tkinter.Entry(self.top2, textvariable=self.top2.item_specs, width=40)
        self.text4.grid(row=4, column=1, columnspan=3)
        self.text5 = tkinter.Entry(self.top2, textvariable=self.top2.blueprint_id, width=40)
        self.text5.grid(row=5, column=1, columnspan=3)
        self.text6 = tkinter.Entry(self.top2, textvariable=self.top2.drawing_id, width=55)
        self.text6.grid(row=1, column=6, columnspan=4)
        self.text7 = tkinter.Entry(self.top2, textvariable=self.top2.drawing_name, width=55)
        self.text7.grid(row=2, column=6, columnspan=4)
        self.text8 = tkinter.Entry(self.top2, textvariable=self.top2.drawing_code, width=55)
        self.text8.grid(row=3, column=6, columnspan=4)
        self.text9 = tkinter.Entry(self.top2, textvariable=self.top2.drawing_identifier, width=55)
        self.text9.grid(row=4, column=6, columnspan=4)
        self.text10 = tkinter.Entry(self.top2, textvariable=self.top2.drawing_parameter, width=55)
        self.text10.grid(row=5, column=6, columnspan=4)
        self.text11 = tkinter.Entry(self.top2, textvariable=self.top2.drawing_edition, width=55)
        self.text11.grid(row=6, column=6, columnspan=4)
        #self.text12 = tkinter.Entry(self.top2, textvariable=self.top2.drawing_classification, width=55)
        self.text12=tkinter.ttk.Combobox(self.top2,textvariable=self.top2.drawing_classification,width=52,values=classify_list)
        self.text12.grid(row=7, column=6, columnspan=4)
        self.text13 = tkinter.Entry(self.top2, textvariable=self.top2.orderDate, width=55)
        self.text13.grid(row=8, column=6, columnspan=4)
        self.text14 = tkinter.Entry(self.top2, textvariable=self.top2.document_name, width=55)
        self.text14.grid(row=9, column=6, columnspan=4)
        self.text15 = tkinter.Entry(self.top2, textvariable=self.top2.document_type, width=55)
        self.text15.grid(row=10, column=6, columnspan=4)
        self.text16 = tkinter.Entry(self.top2, textvariable=self.top2.document_path, width=55)
        self.text16.grid(row=11, column=6, columnspan=4)
        self.text17 = tkinter.Entry(self.top2, textvariable=self.top2.drawing_remarks, width=55)
        self.text17.grid(row=12, column=6, columnspan=4)

        self.button1 = tkinter.Button(self.top2, text="新增料号表记录", fg="blue", width=15, command=self.insertItem).grid(row=13, column=1)
        self.button1 = tkinter.Button(self.top2, text="新增图纸图号表", fg="blue", width=15,command=self.insertDrawing).grid(row=13, column=6)
        self.button1 = tkinter.Button(self.top2, text="清空料号表文本框", fg="red", width=15, command=self.cleaitemNo).grid(row=13, column=2)
        self.button1 = tkinter.Button(self.top2, text="清空图号表文本框", fg="red", width=15, command=self.cleadrawingNo).grid(row=13, column=7)
        self.button1 = tkinter.Button(self.top2, text="退  出", fg="black", width=15, command=self.top2quit).grid(row=13,column=9)

        def dragged_files(files):   # 获取文件路径和文件名及扩展名
            fullpath = "\n".join((item.decode("gbk") for item in files)).replace("\\","/")
            filetype=fullpath[len(fullpath)-fullpath[::-1].index("."):]
            filename=fullpath[len(fullpath)-fullpath[::-1].index("/"):len(fullpath)-fullpath[::-1].index(".")-1]
            new_path = fullpath.replace(main_path, "")
            self.top2.document_type.set(value=filetype)
            self.top2.document_path.set(value=fullpath)
            self.top2.document_name.set(value=filename)
            self.top2.drawing_name.set(value=filename)
            self.top2.drawing_classification.set(value=new_path[:new_path.index("/")])

        def calendar(event):  # 日历选择小部件
            def print_sel():
                self.text13.delete(0, tkinter.END)
                self.text13.insert(0, str(cal.selection_get()))
                top_cal.destroy()

            x, y = event.widget.winfo_pointerxy()  # 获取鼠标坐标
            top_cal = tkinter.Toplevel(self.top2)
            w, h = top_cal.winfo_screenwidth(), top_cal.winfo_screenheight()  # 获取屏幕分辨率的长和宽
            if x + 250 > w:  # 如果鼠标X位置+小部件宽大于屏幕宽
                x = w - 260
            if y + 200 > h-80:
                y = h - 280
            top_cal.geometry(f"250x200+{str(x)}+{str(y)}")
            top_cal.attributes("-topmost", True)  # 窗口始终置项
            cal = Calendar(top_cal, locale="zh", selectmode='day', cursor="hand1",year=datetime.now().year, month=datetime.now().month,day=datetime.now().day)
            cal.pack(fill="both", expand=True)
            tkinter.Button(top_cal, text=" 确 定 ", command=print_sel).pack()

        self.button_cal = tkinter.Button(self.top2, text='v', command=calendar, width=1,bd=0)
        self.button_cal.grid(row=8, column=9, sticky="e")
        self.button_cal.bind("<Button-1>", calendar)

        windnd.hook_dropfiles(self.Lable20,func=dragged_files)

        self.top2.protocol("WM_DELETE_WINDOW", self.top2quit)  # 点击关闭X实现关闭并返回主窗口

    def insertItem(self):
        #print(self.text1.get(),self.text2.get(),self.text3.get(),self.text4.get(),self.text5.get())
        repeat_list=self.drawingId_query(drawingId=self.text5.get())
        #print(repeat_list,len(repeat_list))
        if len(repeat_list)==0 or self.text2.get()=="":
            tkinter.messagebox.showerror(title="错误",message="料号和品名必填，添加料号前应先添加对应的图纸资料再添加料号相关资料！")
        else:
            insert_sql=f"insert into partnumber (item_id,item_No,item_name,item_specs,blueprint_id) values ({self.text1.get()},'{self.text2.get()}','{self.text3.get()}','{self.text4.get()}',{self.text5.get()})"
            #print("插入语句：",insert_sql)
            conn = pymysql.connect(host=servername, user=serveruser, password=serverpassword, database=serverdatabase)
            cursor_6 = conn.cursor()
            cursor_6.execute(insert_sql)
            #item_list = cursor_5.fetchall()
            conn.commit()
            cursor_6.close()
            conn.close()
            tkinter.messagebox.showinfo(parent=self.top2,title="成功", message="料号表记录添加成功！")

    def insertDrawing(self):
        repeat_list=self.drawingId_query(drawingId=self.text6.get())
        drawinglist=f"({self.text6.get()},'{self.text7.get()}','{self.text8.get()}','{self.text9.get()}','{self.text10.get()}','{self.text11.get()}','{self.text12.get()}','{self.text13.get()}','{self.text14.get()}','{self.text15.get()}','{self.text16.get()}','{self.text17.get()}')"
        if len(repeat_list)!=0 or self.text7.get()=="":
            tkinter.messagebox.showerror(title="错误",message="图号重复或图纸名称未填！")
        else:
            insert_sql=f"insert into blueprint (drawing_id,drawing_name,drawing_code,drawing_identifier,drawing_parameter,drawing_edition,drawing_classification,orderDate,document_name,document_type,document_path,drawing_remarks) values {drawinglist}"
            #print("插入语句：",insert_sql)
            conn = pymysql.connect(host=servername, user=serveruser, password=serverpassword, database=serverdatabase)
            cursor_7 = conn.cursor()
            cursor_7.execute(insert_sql)
            conn.commit()
            cursor_7.close()
            conn.close()
            tkinter.messagebox.showinfo(parent=self.top2,title="成功", message="图纸图号记录添加成功！")

    def batchInsert(self):
        self.top3 = tkinter.Toplevel()
        # self.master1.state("iconic")    # 最小化上级窗口
        self.top3.title("从Excel模板文件批量新增")
        screenheight = self.top3.winfo_screenheight() / 2
        screenwidth = self.top3.winfo_screenwidth() / 2
        tkwidth = 700
        tkheight = 523
        self.top3.geometry(f"{tkwidth}x{tkheight}")
        self.top3.geometry(f"+{int(screenwidth - tkwidth / 2)}+{int(screenheight - tkheight / 2)}")

        imagefilename = tkinter.PhotoImage(file="背景1.gif")
        self.background_label = tkinter.Label(self.top3, text="选择Excel模板分批导入，导入料号表前需要先导入关联的图纸图号表", font=("黑体", 16), image=imagefilename,compound="bottom",background="#f0f0f0", foreground="red")
        self.background_label.place(x=0, y=0, relwidth=1, relheight=1)
        self.background_label.image = imagefilename
        self.flash()

        self.button1 = tkinter.Button(self.top3, text="料号表批量新增", bg="#f0f0f0", font=("宋体", 10),command=self.batchinsertItem)
        self.button1.place(x=275, y=200, width=150, height=40)
        self.button3 = tkinter.Button(self.top3, text="图纸图号表批量新增", bg="#f0fff0", font=("宋体", 10),command=self.batchinsertDrawing)
        self.button3.place(x=275, y=250, width=150, height=40)
        self.button2 = tkinter.Button(self.top3, text="退  出", bg="#f0f0f0", font=("宋体", 18), command=self.top3quit)
        self.button2.place(x=275, y=300, width=150, height=40)

    def batchinsertItem(self):
        file_path = filedialog.askopenfilename(title="选择Excel文件", filetypes=(("xlsx", "*.xlsx"), ("xls", "*.xls"),))
        try:
            wb=openpyxl.load_workbook(file_path)
            if wb.sheetnames[0].title()=="料号表":
                ws=wb.worksheets[0]
                conn = pymysql.connect(host=servername, user=serveruser, password=serverpassword,database=serverdatabase)
                for row in range(ws.max_row-1):
                    row_list=[]
                    for col in range(5):
                        if ws.cell(row+2,col+1).value==None:
                            ws.cell(row+2,col+1).value=""
                        row_list.append(ws.cell(row+2,col+1).value)
                    drawingsql=f"select drawing_name from blueprint where drawing_id={row_list[4]}"
                    sql=f"insert into partnumber (item_id,item_No,item_name,item_specs,blueprint_id) values ({row_list[0]},'{row_list[1]}','{row_list[2]}','{row_list[3]}',{row_list[4]})"
                    cursor_9 = conn.cursor()
                    cursor_9.execute(drawingsql)
                    item_list = cursor_9.fetchall()
                    cursor_9.close()
                    if len(item_list)==0:
                        tkinter.messagebox.showerror(title="导入错误",message=f"第{row+1}行图号{row_list[4]}不存在，请先添加对应图纸图号信息！")
                        continue
                    cursor_10 = conn.cursor()
                    cursor_10.execute(sql)
                    conn.commit()
                    cursor_10.close()
                conn.close()
                wb.close()
                tkinter.messagebox.showinfo(title="导入完成", message="成功导入料号资料。")
            else:
                tkinter.messagebox.showerror(title="错误", message="打开的文件不是批量导入模板文件。")
                wb.close()
        except:
            tkinter.messagebox.showwarning(title="提示", message="未选择批量导入模板文件。")

    def batchinsertDrawing(self):
        file_path = filedialog.askopenfilename(title="选择Excel文件", filetypes=(("xlsx", "*.xlsx"), ("xls", "*.xls"),))
        try:
            wb = openpyxl.load_workbook(file_path)
            if wb.sheetnames[0].title() == "料号表":
                ws = wb.worksheets[1]
                conn = pymysql.connect(host=servername, user=serveruser, password=serverpassword, database=serverdatabase)
                for row in range(ws.max_row - 1):
                    row_list = []
                    for col in range(12):
                        if ws.cell(row+2,col+1).value==None:
                            ws.cell(row+2,col+1).value=""
                        row_list.append(ws.cell(row + 2, col + 1).value)
                    drawingsql = f"select drawing_id from blueprint where drawing_id={row_list[0]}"
                    sql = f"insert into blueprint (drawing_id,drawing_name,drawing_code,drawing_identifier,drawing_parameter,drawing_edition,drawing_classification,orderDate,document_name,document_type,document_path,drawing_remarks) values ({row_list[0]},'{row_list[1]}','{row_list[2]}','{row_list[3]}','{row_list[4]}','{row_list[5]}','{row_list[6]}','{row_list[7]}','{row_list[8]}','{row_list[9]}','{row_list[10]}','{row_list[11]}')"
                    cursor_11 = conn.cursor()
                    cursor_11.execute(drawingsql)
                    item_list = cursor_11.fetchall()
                    cursor_11.close()
                    if len(item_list) != 0:
                        tkinter.messagebox.showinfo(title="导入错误", message=f"第{row + 1}行图号{row_list[4]}已经存在！")
                        continue
                    cursor_12 = conn.cursor()
                    cursor_12.execute(sql)
                    conn.commit()
                    cursor_12.close()
                conn.close()
                wb.close()
                tkinter.messagebox.showinfo(title="导入完成", message="成功导入图纸图号资料。")
            else:
                tkinter.messagebox.showinfo(title="错误", message="打开的文件不是批量导入模板文件。")
                wb.close()
        except:
            tkinter.messagebox.showinfo(title="提示", message="未选择批量导入模板文件。")

    def flash(self):    # 标签闪烁提醒
        bg = self.background_label.cget("background")
        fg = self.background_label.cget("foreground")
        self.background_label.configure(background=fg, foreground=bg)
        self.background_label.after(300, self.flash)

    def modifyItem(self):
        if self.text1.get()=="" or self.text5.get()=="":
            tkinter.messagebox.showerror(title="错误", message="编号和图纸ID为必填项，不能为空。")
        else:
            item_id=int(self.text1.get())
            blueprint_id=int(self.text5.get())
            itemvalues=[item_id,self.text2.get(),self.text3.get(),self.text4.get(),blueprint_id]
            sql=f"update partnumber set item_No='{itemvalues[1]}',item_name='{itemvalues[2]}',item_specs='{itemvalues[3]}',blueprint_id={itemvalues[4]} where item_id={item_id}"
            conn = pymysql.connect(host=servername, user=serveruser, password=serverpassword, database=serverdatabase)
            cursor_3 = conn.cursor()
            cursor_3.execute(sql)
            conn.commit()
            cursor_3.close()
            conn.close()
            tkinter.messagebox.showinfo(parent=self.top1,title="成功", message="修改成功！")

    def modifyblueprint(self):
        if self.text5.get()=="" or self.text6.get()=="":
            tkinter.messagebox.showerror(title="错误", message="图纸ID和图纸名称为必填项，不能为空。")
        else:
            drawing_id=int(self.text5.get())
        drawingvalues=[self.text7.get(),self.text8.get(),self.text9.get(),self.text10.get(),self.text11.get(),self.text12.get(),self.text13.get(),self.text14.get(),self.text15.get(),self.text16.get(),self.text17.get()]
        sql=f"update blueprint set drawing_name='{drawingvalues[0]}',drawing_code='{drawingvalues[1]}',drawing_identifier='{drawingvalues[2]}',drawing_parameter='{drawingvalues[3]}',drawing_edition='{drawingvalues[4]}',drawing_classification='{drawingvalues[5]}',orderDate='{drawingvalues[6]}',document_name='{drawingvalues[7]}',document_type='{drawingvalues[8]}',document_path='{drawingvalues[9]}',drawing_remarks='{drawingvalues[10]}' where drawing_id={drawing_id}"
        conn = pymysql.connect(host=servername, user=serveruser, password=serverpassword, database=serverdatabase)
        cursor_3 = conn.cursor()
        cursor_3.execute(sql)
        conn.commit()
        cursor_3.close()
        conn.close()
        tkinter.messagebox.showinfo(parent=self.top1,title="成功", message="修改成功！")

    def cleaitemNo(self):
        self.text1.delete(0,tkinter.END)
        self.text2.delete(0, tkinter.END)
        self.text3.delete(0, tkinter.END)
        self.text4.delete(0, tkinter.END)
        self.text5.delete(0, tkinter.END)

    def cleadrawingNo(self):
        self.text6.delete(0,tkinter.END)
        self.text7.delete(0, tkinter.END)
        self.text8.delete(0, tkinter.END)
        self.text9.delete(0, tkinter.END)
        self.text10.delete(0, tkinter.END)
        self.text11.delete(0, tkinter.END)
        self.text12.delete(0, tkinter.END)
        self.text13.delete(0, tkinter.END)
        self.text14.delete(0, tkinter.END)
        self.text15.delete(0, tkinter.END)
        self.text16.delete(0, tkinter.END)
        self.text17.delete(0, tkinter.END)

    def top1quit(self):
        self.top1.destroy()  # 关闭当前窗口

    def top2quit(self):
        self.top2.destroy()  # 关闭当前窗口
        self.master1.state("normal")

    def top3quit(self):
        self.top3.destroy()  # 关闭当前窗口

    def DrawingNoquit(self):
        self.master1.quit()

    def deleteItem(self):
        confirm=tkinter.messagebox.askokcancel(title="删除确认",message="确认需要删除选中项么？")
        if confirm==True:
            curItem = self.data_table.selection()
            #print("选中行是", curItem,len(curItem),self.data_table.item(curItem[0])["values"][0])

            if len(curItem)>0:
                conn = pymysql.connect(host=servername, user=serveruser, password=serverpassword,
                                       database=serverdatabase)
                num=0
                for i in range(len(curItem)):
                    if self.data_table.item(curItem[i])['values'][0] == "":
                        tkinter.messagebox.showwarning(title="错误", message="无记录！")
                        break
                    sql=f"delete from partnumber where item_id={self.data_table.item(curItem[i])['values'][0]}"
                    #print(sql)
                    cursor_8 = conn.cursor()
                    cursor_8.execute(sql)
                    conn.commit()
                    cursor_8.close()
                    num+=1
                conn.close()
                tkinter.messagebox.showinfo(title="成功", message=f"已删除{num}条记录！")
            else:
                tkinter.messagebox.showerror(title="错误",message="未选择数据！")
        # else:
        #     print("取消删除！")
        #     pass

    def deleteDrawing(self):
        confirm=tkinter.messagebox.askokcancel(title="删除确认",message="确认需要删除选中项么？")
        if confirm==True:
            curItem = self.data_table.selection()
            #print("选中行是", curItem,len(curItem),self.data_table.item(curItem[0])["values"][0],self.data_table.item(curItem[0])["values"][4])
            if len(curItem)>0:
                num=0
                conn = pymysql.connect(host=servername, user=serveruser, password=serverpassword,
                                       database=serverdatabase)
                for i in range(len(curItem)):
                    sql=f"delete from blueprint where drawing_id={self.data_table.item(curItem[i])['values'][4]}"
                    itemsql=f"select item_id,blueprint_id from partnumber where blueprint_id={self.data_table.item(curItem[i])['values'][4]}"
                    #print(sql,"\n",itemsql)
                    cursor=conn.cursor()
                    cursor.execute(itemsql)
                    item_list=cursor.fetchall()
                    cursor.close()
                    #print(item_list,len(item_list))
                    if len(item_list)>0:
                        tkinter.messagebox.showerror(title="错误", message=f"料号表中有关联图纸ID为{self.data_table.item(curItem[i])['values'][4]}的记录！")
                        break
                    else:
                        cursor_8 = conn.cursor()
                        cursor_8.execute(sql)
                        conn.commit()
                        cursor_8.close()
                        num+=1
                conn.close()
                tkinter.messagebox.showinfo(title="成功", message=f"已删除{num}条记录！")
            else:
                tkinter.messagebox.showerror(title="错误",message="未选择数据！")
        # else:
        #     print("取消删除！")
        #     pass

    def exportExcel(self):
        titles = ["ID", "料号", "品名", "规格", "图纸ID", "图纸名称", "图纸编码", "图纸编号", "图纸参数", "图纸版本", "图纸分类", "创建时间", "文件名称",
                  "文件类型", "文件路径", "图纸备注"]
        now=datetime.now()
        nowtime=str(now.date()).replace("-","")+str(now.strftime("%H%M%S"))
        excelfilename=export_path+"图纸图号管理导出"+nowtime+".xlsx"
        wb=openpyxl.Workbook()
        ws=wb.active
        ws.append(titles)
        for row_id in self.data_table.get_children():
            row=self.data_table.item(row_id)["text"].replace("'","\"")
            # print(type(row))
            # print(row)
            # print(eval(row))
            #str_row=str("[\"")+str(row).replace(" ","\",\"").replace("{}","")+str("\"]")
            #print(str_row)
            #ws.append(literal_eval(str_row))
            ws.append(eval(row))

        wb.save(excelfilename)
        wb.close()
        tkinter.messagebox.showinfo(title="导出成功", message=f"成功导出到文件{excelfilename}")

def main():

    root=tkinter.Tk()
    myGUIWelcome=Welcome(root)

    root.mainloop()

if __name__ == '__main__':
  main()